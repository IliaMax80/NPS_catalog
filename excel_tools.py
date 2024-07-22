import os
import pathlib
import sys


from openpyxl import load_workbook, Workbook
from datetime import date, datetime

import locale

DEFOLT_INFORMATION_FIELD_EXCEL = {'фио': 'text', 'должность': 'text', 'дата приема': 'date',
                                  'дата рождения': 'date', 'серия': 'text', 'номер': 'text',
                                  'паспорт дата выдачи': 'date', 'паспорт выдан': 'text', 'код подразделения': 'text',
                                  'адрес регестрации': 'text', 'снилс': 'text', 'инн': 'text'}
MONTH = {'января': 'январь', 'февраля': 'февраль', 'марта': 'март', 'апреля': 'апрель',
         'мая': 'май', 'июня': 'июнь', 'июля': 'июль', 'августа': 'август', 'сентября': 'сентябрь',
         'октября': 'октябрь', 'ноября': 'ноябрь', 'декабря': 'декабрь'}


def import_table_worker(workers, file):
    workers.add_category('водитель')
    workers.add_sample_information_fields_category('водитель', DEFOLT_INFORMATION_FIELD_EXCEL)
    workers.add_category('рабочий')
    workers.add_sample_information_fields_category('рабочий', DEFOLT_INFORMATION_FIELD_EXCEL)
    workbook = load_workbook(file)
    worksheet = workbook.active
    count = int(worksheet.cell(2, 5).value)
    start_row = 7
    for row in range(start_row, start_row + count):
        information_fields = {}
        id = int(worksheet.cell(row, 2).value)
        full_name = worksheet.cell(row, 4).value
        post = worksheet.cell(row, 8).value.lower()
        information_fields['фио'] = full_name
        information_fields['должность'] = post
        information_fields['дата приема'] = worksheet.cell(row, 16).value
        information_fields['дата рождения'] = worksheet.cell(row, 17).value
        information_fields = str_to_passport(worksheet.cell(row, 18).value, information_fields)
        information_fields['адрес регестрации'] = worksheet.cell(row, 19).value
        information_fields['снилс'] = worksheet.cell(row, 20).value
        information_fields['инн'] = worksheet.cell(row, 21).value
        category = 'водитель' if post == 'водитель автомобиля' else 'рабочий'
        document = {}
        workers.append_data_legal_unit(category, id, information_fields, document)
    workers.save()


def str_to_passport(string, information_fields):
    locale.setlocale(locale.LC_ALL, '')
    strings = [i.lower().strip() for i in string.split(', ')]
    if len(strings) != 6:
        return None
    information_fields['серия'] = ''.join([i for i in strings[1] if i.isdigit()])
    information_fields['номер'] = ''.join([i for i in strings[2] if i.isdigit()])
    date_words = strings[3].split()
    date_words[2] = MONTH[date_words[2]]
    date_str = f'{date_words[1]} {date_words[2]} {date_words[3]}'
    dt = datetime.strptime(date_str, '%d %B %Y')
    information_fields['паспорт дата выдачи'] = date(dt.year, dt.month, dt.day).strftime('%d.%m.%Y')
    information_fields['паспорт выдан'] = strings[4]
    information_fields['код подразделения'] = ''.join([i for i in strings[5] if i.isdigit() or i == '-'])
    return information_fields

def export_table_to_excel(file, model_table):
    print(file)
    if file[-5:] != '.xlsx':
        file = file + '.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    for column in range(0, len(model_table.columns)):
        worksheet.cell(1, column + 2, model_table.columns[column])
    for row in range(len(model_table.id_legal_units)):
        worksheet.cell(row + 2, 1, row + 1)
        for column in range(len(model_table.columns)):
            worksheet.cell(row + 2, column + 2, model_table.item(row, column).text())
    workbook.save(file)